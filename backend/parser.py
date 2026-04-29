import re
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io


def parse_nsdl_pdf(pdf_path):
    client_name = None
    current_security = None
    pending_opening = None
    records = []
    last_bal = {}

    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                full_text += t + "\n"

    lines = full_text.split("\n")

    for line in lines:
        # Client name
        if client_name is None:
            m = re.match(r'^Name\s+([A-Z][A-Z ]+?)$', line.strip())
            if m:
                client_name = m.group(1).strip()

        # ISIN + Security
        isin_m = re.match(r'^\s*ISIN\s+(INE\w+)\s+(.+)', line)
        if isin_m:
            current_security = isin_m.group(2).strip()
            pending_opening = None
            continue

        # Opening balance
        ob_m = re.search(r'Opening Balance\s*:\s*([\d,]+)', line)
        if ob_m and current_security:
            pending_opening = int(ob_m.group(1).replace(',', ''))
            continue

        # Transaction row
        tx_m = re.match(r'^\s*(\d{2}-[A-Za-z]{3}-\d{4})\s+(\d{9,})\s+(.+)', line)
        if tx_m and current_security:
            date, txno, rest = tx_m.groups()
            rest = rest.strip()
            is_credit = rest.startswith('By ')
            is_debit = rest.startswith('To ')

            nums = [int(n.replace(',', '')) for n in re.findall(r'\b[\d,]+\b', rest)
                    if n.replace(',', '').isdigit() and n.replace(',', '')]
            if len(nums) < 2:
                continue

            closing = nums[-1]
            qty = nums[-2]

            if pending_opening is not None:
                open_bal = pending_opening
                pending_opening = None
            elif current_security in last_bal:
                open_bal = last_bal[current_security]
            else:
                open_bal = (closing - qty) if is_credit else (closing + qty)

            last_bal[current_security] = closing

            records.append({
                'client_name': client_name or '',
                'security_name': current_security,
                'transaction_date': date,
                'transaction_number': txno,
                'opening_balance': open_bal,
                'credit': qty if is_credit else 0,
                'debit': qty if is_debit else 0,
                'closing_balance': closing,
            })

    return records, client_name


def records_to_excel(records, client_name=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    header_fill = PatternFill("solid", fgColor="1B3A6B")
    alt_fill = PatternFill("solid", fgColor="EBF0FA")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    credit_fill = PatternFill("solid", fgColor="E8FFF4")
    debit_fill = PatternFill("solid", fgColor="FFF0F0")

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Client Name", "Security Name", "Transaction Date",
               "Transaction Number", "Opening Balance", "Credit",
               "Debit", "Closing Balance"]
    col_widths = [28, 44, 18, 22, 18, 14, 14, 18]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 22

    for row_idx, r in enumerate(records, 2):
        if r['credit'] > 0:
            rf = credit_fill
        elif r['debit'] > 0:
            rf = debit_fill
        else:
            rf = alt_fill if row_idx % 2 == 0 else white_fill

        values = [
            r['client_name'], r['security_name'], r['transaction_date'],
            r['transaction_number'], r['opening_balance'],
            r['credit'] if r['credit'] > 0 else '',
            r['debit'] if r['debit'] > 0 else '',
            r['closing_balance'],
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.fill = rf
            cell.border = border
            if col in (5, 8) and isinstance(value, int):
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0'
            elif col in (6, 7) and isinstance(value, int):
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0'
                cell.font = Font(name="Arial", size=10,
                                 color="107954" if col == 6 else "C0392B", bold=True)
            elif col == 3:
                cell.alignment = Alignment(horizontal='center')
            elif col == 4:
                cell.alignment = Alignment(horizontal='left')
                cell.font = Font(name="Courier New", size=9)
            else:
                cell.alignment = Alignment(horizontal='left')

    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Summary")
    ws2['A1'] = "NSDL Transaction Statement Summary"
    ws2['A1'].font = Font(bold=True, size=14, name="Arial", color="1B3A6B")
    summary_data = [
        ("Client Name", client_name or "N/A"),
        ("Total Transactions", len(records)),
        ("Unique Securities", len(set(r['security_name'] for r in records))),
        ("Total Credits (qty)", sum(r['credit'] for r in records)),
        ("Total Debits (qty)", sum(r['debit'] for r in records)),
    ]
    for i, (label, value) in enumerate(summary_data, 3):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True, name="Arial")
        ws2.cell(row=i, column=2, value=value).font = Font(name="Arial")
    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
